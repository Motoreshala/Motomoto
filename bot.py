import os
print("✅ Я Бот! Запущен вот из этого файла:", __file__)
print("🕒 Версия от 28.05.2025 21:99")
print("🖥️  Текущий PID:", os.getpid())
import logging
from datetime import datetime
from functools import wraps
from time import time

import pandas as pd
import gspread
from gspread.utils import rowcol_to_a1
from oauth2client.service_account import ServiceAccountCredentials

from telegram import (
    Update, InputFile,
    InlineKeyboardMarkup, InlineKeyboardButton,
    ReplyKeyboardMarkup,
)
from telegram.ext import (
    Updater, CommandHandler, MessageHandler, CallbackQueryHandler,
    Filters as TgFilters, CallbackContext
)

# Логирование
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO
)
logger = logging.getLogger("bot")

# Константы
QUEUE_TIMEOUT_SEC = 60
SPREADSHEET_KEY = "1Zivk7VGdmHxufb93uX6mXHnG6mCrQI6O5e509mYTQx0"

ALLOWED_ROLLBACK_USER_ID = 470225477

SHEET_STOCK_NAME = "Наш склад"
SHEET_ORDERS_NAME = "Заказы"
SHEET_MOVE_NAME = "Перемещение"
SHEET_MAIN_NAME = "База"
SHEET_TM_NAME = "Заказ ТМ"
SHEET_STOCK_HISTORY = "История остатков"
SHEET_KITS_NAME = "Комплекты"

MOVE_COLUMNS = [
    "Название", "Артикул", "Комплект", "Кол-во",
    "Мп", "Остатки", "Ячейка"
]

TECHNOMARIN_BRANDS = {"Техномарин", "Easterner", "SeaFlo", "Solas", "SPI", "Sumar"}

# ——— ПОДКЛЮЧЕНИЕ К GOOGLE SHEETS ————————————————————
scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
creds = ServiceAccountCredentials.from_json_keyfile_name("service_account.json", scope)
gc = gspread.authorize(creds)
sh = gc.open_by_key(SPREADSHEET_KEY)

sheet_stock  = sh.worksheet(SHEET_STOCK_NAME)
sheet_orders = sh.worksheet(SHEET_ORDERS_NAME)
sheet_main   = sh.worksheet(SHEET_MAIN_NAME)
sheet_kits = sh.worksheet("Комплекты")
sheet_kits = sh.worksheet(SHEET_KITS_NAME)
sheet_tm     = sh.worksheet(SHEET_TM_NAME)

# ——— УТИЛИТЫ ————————————————————
def to_int(x, default=0):
    try:
        return int(float(x))
    except:
        return default

def is_private_chat(update: Update) -> bool:
    return update.effective_chat and update.effective_chat.type == "private"

def get_menu_keyboard(user_id, menu="main"):
    """
    Возвращает разные раскладки кнопок в зависимости от текущего меню.
    menu == "main"     — главная: Сборка / Остатки (+ скрытая Откат для админа)
    menu == "assembly" — кнопки загрузки и обработки заказов + «Назад»
    menu == "stock"    — кнопки управления остатками + «Назад»
    """
    if menu == "main":
        buttons = [["📦 Сборка", "📊 Остатки"]]
        if user_id == ALLOWED_ROLLBACK_USER_ID:
            buttons.append(["♻️ Откатить остатки"])
        return ReplyKeyboardMarkup(buttons, resize_keyboard=True)

    if menu == "assembly":
        buttons = [
            ["📥 Загрузить заказы WB", "📥 Загрузить заказы Ozon"],
            ["📦 Обработать заказы"],
            ["🔙 Назад"]
        ]
        return ReplyKeyboardMarkup(buttons, resize_keyboard=True)

    if menu == "stock":
        buttons = [
            ["➕ Прибавить остатки", "➖ Вычесть остатки"],
            ["✏️ Ручное прибавление", "✏️ Ручное вычитание"],
            ["🔙 Назад"]
        ]
        return ReplyKeyboardMarkup(buttons, resize_keyboard=True)

    # fallback
    return ReplyKeyboardMarkup([["🔙 Назад"]], resize_keyboard=True)

# ——— ОЧЕРЕДЬ (Блокировка на время выполнения) ————————————————————
def queue_guard(func):
        @wraps(func)
        def wrapper(update: Update, context: CallbackContext, *args, **kwargs):
            now = time()
            busy = context.bot_data.get("busy", False)
            ts = context.bot_data.get("busy_ts", 0)

            print(f"🟡 Вызов {func.__name__}, busy={busy}, ts_diff={now - ts}", flush=True)

            if busy and now - ts < QUEUE_TIMEOUT_SEC:
                if update.message:
                    update.message.reply_text("⏳ Подождите, идёт выполнение предыдущей операции.")
                elif update.callback_query:
                    update.callback_query.answer("⏳ Подождите, идёт выполнение предыдущей операции.", show_alert=True)
                return

            context.bot_data["busy"] = True
            context.bot_data["busy_ts"] = now
            try:
                return func(update, context, *args, **kwargs)
            finally:
                context.bot_data["busy"] = False
                context.bot_data["busy_ts"] = 0
        return wrapper

def handle_message(update: Update, context: CallbackContext):
    if not update.message or not update.message.text:
        return

    txt = update.message.text.strip().lower()
    menu = context.user_data.get("menu", "main")

    if txt == "🚀 обработать заказы":
        if context.user_data.get("menu") in ["after_upload_wb", "after_upload_ozon"]:
            process_orders(update, context)
            context.user_data["menu"] = "main"
            return update.message.reply_text(
                "Готово. Главное меню",
                reply_markup=get_menu_keyboard(update.effective_user.id, menu="main")
            )
        else:
            return update.message.reply_text(
                "⚠️ Сначала загрузите заказы WB и Ozon.",
                reply_markup=get_menu_keyboard(update.effective_user.id, menu="main")
            )
    if context.user_data.get("action") in ("manual_set", "manual_sub"):
        action = context.user_data["action"]
        del context.user_data["action"]

        txt_lines = update.message.text.strip().splitlines()
        parsed = []
        for line in txt_lines:
            parts = line.strip().split()
            if len(parts) == 2:
                art, qty = parts
                parsed.append((art.strip(), to_int(qty)))

        if not parsed:
            return update.message.reply_text("⚠️ Формат: Артикул Количество (в одной строке)")

        vals = sheet_stock.get_all_values()
        hdr = vals[0]
        art_i, qty_i = hdr.index("Артикул"), hdr.index("Остатки склад")
        stock_map = {
            str(r[art_i]).strip(): {"row": idx + 2, "stock": to_int(r[qty_i])}
            for idx, r in enumerate(vals[1:])
        }

        report = []
        for art, q in parsed:
            rec = {"Артикул": art, "Было": "-", "+/-": q, "Стало": "-", "Примечание": ""}
            if art in stock_map:
                old = stock_map[art]["stock"]
                new = old + q if action == "manual_set" else max(0, old - q)
                if old - q < 0 and action == "manual_sub":
                    rec["Примечание"] = "недостаток, обнулено"
                stock_map[art]["stock"] = new
                rec["Было"], rec["Стало"] = old, new
            else:
                rec["Примечание"] = "артикул не найден"
            report.append(rec)

        updates = [
            {"range": rowcol_to_a1(info["row"], qty_i + 1), "values": [[info["stock"]]]}
            for art, info in stock_map.items()
        ]
        if updates:
            sheet_stock.batch_update(updates)

        df_r = pd.DataFrame(report, columns=["Артикул", "Было", "+/-", "Стало", "Примечание"])
        username = update.effective_user.full_name or update.effective_user.username
        ts = datetime.now().strftime("%d.%m.%Y %H:%M")
        save_and_send_excel(
            df_r,
            prefix=f"report_manual_{action}",
            caption=f"✏️ Ручное {'прибавление' if action == 'manual_set' else 'вычитание'}\n👤 {username}\n🕓 {ts}",
            update=update,
            context=context,
        )
        return
    # — Главный экран —
    if menu == "main":
        if txt == "📦 сборка":
            context.user_data["menu"] = "assembly"
            return update.message.reply_text(
                "Меню сборки:",
                reply_markup=get_menu_keyboard(update.effective_user.id, menu="assembly")
            )

        if txt == "📊 остатки":
            context.user_data["menu"] = "stock"
            return update.message.reply_text(
                "Меню остатков:",
                reply_markup=get_menu_keyboard(update.effective_user.id, menu="stock")
            )

        if txt == "♻️ откатить остатки" and update.effective_user.id == ALLOWED_ROLLBACK_USER_ID:
            return rollback_stock(update, context)

        # всё прочее на главном — показываем помощь
        return show_help(update, context)

    # — Меню «Сборка» —
    if menu == "assembly":
        if txt == "📦 сборка":
            return update.message.reply_text(
                "Меню сборки:",
                reply_markup=get_menu_keyboard(update.effective_user.id, menu="assembly")
            )

        if txt == "📥 загрузить заказы wb":
            context.user_data["action"] = "load_wb"
            return update.message.reply_text(
                "📥 Пришлите файл WB",
                reply_markup=get_menu_keyboard(update.effective_user.id, menu="assembly")
            )

        if txt == "📥 загрузить заказы ozon":
            context.user_data["action"] = "load_ozon"
            return update.message.reply_text(
                "📥 Пришлите файл Ozon",
                reply_markup=get_menu_keyboard(update.effective_user.id, menu="assembly")
            )

        if txt == "📦 обработать заказы":
            context.user_data["action"] = "process_orders"
            return process_orders(update, context)

        if txt == "🔙 назад":
            context.user_data["menu"] = "main"
            return update.message.reply_text(
                "Главное меню:",
                reply_markup=get_menu_keyboard(update.effective_user.id, menu="main")
            )

        return update.message.reply_text(
            "❓ Неизвестная команда в меню сборки.",
            reply_markup=get_menu_keyboard(update.effective_user.id, menu="assembly")
        )
    # — Меню «Остатки» —
    if menu == "stock":
        if txt == "🔙 назад":
            last = context.user_data.pop("last_uploaded_orders", 0)
            if last:
                vals = sheet_orders.get_all_values()
                for _ in range(min(last, max(0, len(vals) - 1))):
                    sheet_orders.delete_rows(len(vals))
                    vals.pop()
                update.message.reply_text("↩️ Загрузка отменена, заказы удалены.")
            else:
                update.message.reply_text("↩️ Возврат без удаления.")

            context.user_data["menu"] = "main"
            return update.message.reply_text(
                "Главное меню:",
                reply_markup=get_menu_keyboard(update.effective_user.id, menu="main")
            )

        if txt == "➕ прибавить остатки":
            context.user_data["action"] = "set_excel"
            return update.message.reply_text(
                "📊 Пришлите Excel-файл",
                reply_markup=get_menu_keyboard(update.effective_user.id, menu="stock")
            )

        if txt == "➖ вычесть остатки":
            context.user_data["action"] = "sub_excel"
            return update.message.reply_text(
                "📊 Пришлите Excel-файл",
                reply_markup=get_menu_keyboard(update.effective_user.id, menu="stock")
            )

        if txt == "✏️ ручное прибавление":
            context.user_data["action"] = "manual_set"
            return update.message.reply_text(
                "✏️ Введите: Артикул Количество",
                reply_markup=get_menu_keyboard(update.effective_user.id, menu="stock")
            )

        if txt == "✏️ ручное вычитание":
            context.user_data["action"] = "manual_sub"
            return update.message.reply_text(
                "✏️ Введите: Артикул Количество",
                reply_markup=get_menu_keyboard(update.effective_user.id, menu="stock")
            )

        return update.message.reply_text(
            "❓ Неизвестная команда в меню остатков.",
            reply_markup=get_menu_keyboard(update.effective_user.id, menu="stock")
        )

    # safety net — если что-то пошло не так
    context.user_data["menu"] = "main"
    return update.message.reply_text(
        "Главное меню:",
        reply_markup=get_menu_keyboard(update.effective_user.id, menu="main")
    )

def show_help(update: Update, context: CallbackContext):
    help_text = (
        "*ℹ️ Инструкция по использованию бота:*\n\n"
        "📥 *Загрузка заказов:*\n"
        "1) 📥 Загрузить заказы WB — отправьте Excel-файл заказов Wildberries\n"
        "2) 📥 Загрузить заказы Ozon — отправьте .csv или Excel-файл заказов Ozon\n"
        "3) 📦 Обработать заказы — создаст перемещение, изменит остатки, сформирует заказ ТМ\n\n"
        "📦 *Управление остатками:*\n"
        "1) ➕ / ➖ — загрузите Excel с колонками 'Артикул' и 'Количество'\n"
        "2) ✏️ Ручное прибавление/вычитание — отправьте текст: `артикул количество`\n"
        "3) ♻️ Откатить остатки — доступно только администратору\n"
        "4) 📖 Помощь — показать это меню\n"
        "5) 🔙 Назад — отменить текущее действие\n"
    )
    update.message.reply_text(help_text, parse_mode="Markdown", reply_markup=get_menu_keyboard(update.effective_user.id, menu="assembly"))

# ——— ОБРАБОТКА ДОКУМЕНТОВ ————————————————————
@queue_guard
def handle_document(update: Update, context: CallbackContext):
    if not update.message or not update.message.document:
        return

    action = context.user_data.get("action")
    local = f"/tmp/{datetime.now().timestamp()}"
    file = update.message.document
    ext = file.file_name.split('.')[-1].lower()
    local += f".{ext}"
    file.get_file().download(local)

    try:
        if action in ("load_wb", "load_ozon"):
            if action == "load_ozon":
                try:
                    df = pd.read_csv(local, sep=None, engine="python", encoding="utf-8-sig")
                except:
                    df = pd.read_csv(local, sep=";", encoding="utf-8-sig")
            else:
                df = pd.read_excel(local, engine="openpyxl")

            recs = []
            if action == "load_wb":
                for art in df.get("Артикул продавца", pd.Series()).dropna().astype(str):
                    art = art.strip()
                    # Проверка на дублирование (повтор первой половины)
                    if len(art) % 2 == 0 and art[:len(art)//2] == art[len(art)//2:]:
                        art = art[:len(art)//2]
                    recs.append([art, 1, "WB"])
            else:
                if "Количество" in df.columns:
                    for art, q in zip(df["Артикул"], df["Количество"]):
                        recs.append([str(art).strip(), to_int(q, 1), "Ozon"])
                else:
                    for art in df["Артикул"]:
                        recs.append([str(art).strip(), 1, "Ozon"])

            if recs:
                # ✅ Проверка и установка заголовков
                existing_orders = sheet_orders.get_all_values()
                if not existing_orders or existing_orders[0] != ["Артикул", "Количество", "Маркетплейс"]:
                    sheet_orders.clear()
                    sheet_orders.append_row(["Артикул", "Количество", "Маркетплейс"], value_input_option="USER_ENTERED")

                sheet_orders.append_rows(recs, value_input_option="USER_ENTERED")
                context.user_data["last_uploaded_orders"] = len(recs)
                context.user_data["menu"] = "assembly"
                update.message.reply_text(
                    f"✅ Добавлено {len(recs)} строк в «Заказы»",
                    reply_markup=get_menu_keyboard(update.effective_user.id, menu="assembly")
                )

        elif action in ("set_excel", "sub_excel"):
            context.user_data["menu"] = "stock"
            df = pd.read_excel(local, engine="openpyxl")
            vals = sheet_stock.get_all_values()
            hdr = vals[0]
            art_i, qty_i = hdr.index("Артикул"), hdr.index("Остатки склад")
            stock_map = {
                str(r[art_i]).strip(): {"row": idx + 2, "stock": to_int(r[qty_i])}
                for idx, r in enumerate(vals[1:])
            }
            report = []

            for art, q in zip(df.get("Артикул", []), df.get("Количество", [])):
                art = str(art).strip()
                q = to_int(q)
                rec = {"Артикул": art, "Было": "-", "+/-": q, "Стало": "-", "Примечание": ""}
                if art in stock_map:
                    old = stock_map[art]["stock"]
                    new = old + q if action == "set_excel" else max(0, old - q)
                    if old - q < 0 and action == "sub_excel":
                        rec["Примечание"] = "недостаток, обнулено"
                    stock_map[art]["stock"] = new
                    rec["Было"], rec["Стало"] = old, new
                else:
                    rec["Примечание"] = "артикул не найден"
                report.append(rec)

            updates = [
                {"range": rowcol_to_a1(info["row"], qty_i + 1), "values": [[info["stock"]]]}
                for art, info in stock_map.items()
            ]
            if updates:
                sheet_stock.batch_update(updates)

            df_r = pd.DataFrame(report, columns=["Артикул", "Было", "+/-", "Стало", "Примечание"])
            username = update.effective_user.full_name or update.effective_user.username
            ts = datetime.now().strftime("%d.%m.%Y %H:%M")
            save_and_send_excel(
                df_r,
                prefix=f"report_excel_{action}",
                caption=f"📄 Excel-изменения ({'прибавление' if action == 'set_excel' else 'вычитание'})\n👤 {username}\n🕓 {ts}",
                update=update,
                context=context,
            )
        else:
            update.message.reply_text("⚠️ Сначала выберите действие", reply_markup=get_menu_keyboard(update.effective_user.id, menu="assembly"))

    except Exception as e:
        logger.exception("handle_document")
        update.message.reply_text(f"❌ Ошибка: {e}", reply_markup=get_menu_keyboard(update.effective_user.id, menu="assembly"))
    finally:
        if os.path.exists(local):
            os.remove(local)
        context.user_data.pop("action", None)

# ——— РУЧНОЙ ВВОД ОСТАТКОВ ————————————————————

def _handle_manual(update: Update, context: CallbackContext, mode: str):
    if not update.message or not update.message.text:
        return

    username = update.effective_user.full_name or update.effective_user.username
    lines = update.message.text.strip().splitlines()
    all_vals = sheet_stock.get_all_values()
    hdr = all_vals.pop(0)
    art_i, qty_i = hdr.index("Артикул"), hdr.index("Остатки склад")
    save_stock_snapshot(update.effective_user, reason="manual_input")
    report = []

    for ln in lines:
        try:
            art, q = ln.split()
            q = to_int(q)
        except:
            report.append([ln, "-", "-", "неверный формат"])
            continue

        found = False
        for idx, r in enumerate(all_vals):
            if str(r[art_i]).strip() == art:
                curr = to_int(r[qty_i])
                new = curr + q if mode == "manual_set" else max(0, curr - q)
                note = f"{curr}+{q}={new}" if mode == "manual_set" else (f"{curr}−{q}={new}" if curr >= q else "0 (недостаток)")
                sheet_stock.update_cell(idx + 2, qty_i + 1, new)
                report.append([art, curr, q, new, note])
                found = True
                break
        if not found:
            report.append([art, "-", "-", "не найден"])

    df_r = pd.DataFrame(report, columns=["Артикул", "Было", "+/-", "Стало", "Примечание"])
    fn = f"/tmp/manual_{mode}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    df_r.to_excel(fn, index=False)
    save_and_send_excel(
        df_r, "manual_update",
        f"📄 Ручная операция\n👤 {username}\n🕓 {datetime.now():%d.%m.%Y %H:%M}",
        update, context
    )
    os.remove(fn)
    context.user_data.pop("action", None)

# ——— ОТКАТ ОСТАТКОВ ————————————————————
def rollback_stock(update: Update, context: CallbackContext):
    try:
        sheet_history = sh.worksheet(SHEET_STOCK_HISTORY)
        data = sheet_history.get_all_values()
        if len(data) < 2:
            update.message.reply_text("📭 Нет сохранённых состояний для отката.")
            return

        df = pd.DataFrame(data[1:], columns=data[0])
        recent = df["Дата"].drop_duplicates().sort_values(ascending=False).head(5).tolist()

        keyboard = [[d] for d in recent]
        update.message.reply_text(
            "🕓 Выберите дату для отката остатков:",
            reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
        )
        context.user_data["awaiting_rollback_date"] = True

    except Exception as e:
        update.message.reply_text(f"❌ Ошибка при откате:\n{e}")

def perform_stock_rollback(update: Update, context: CallbackContext):
    if not update.message or not update.message.text:
        return

    context.user_data.pop("awaiting_rollback_date", None)
    chosen_date = update.message.text.strip()

    try:
        sheet_history = sh.worksheet(SHEET_STOCK_HISTORY)
        data = sheet_history.get_all_values()
        df = pd.DataFrame(data[1:], columns=data[0])
        df = df[df["Дата"] == chosen_date]

        if df.empty:
            update.message.reply_text("❌ Нет данных для выбранной даты.")
            return

        stock_vals = sheet_stock.get_all_values()
        hdr = stock_vals[0]
        art_i = hdr.index("Артикул")
        qty_i = hdr.index("Остатки склад")
        current = {row[art_i]: idx + 2 for idx, row in enumerate(stock_vals[1:])}

        updates = []
        for row in df.itertuples(index=False):
            art = getattr(row, "Артикул")
            qty = to_int(getattr(row, "Остаток"))
            if art in current:
                updates.append({
                    "range": rowcol_to_a1(current[art], qty_i + 1),
                    "values": [[qty]]
                })

        if updates:
            sheet_stock.batch_update(updates)
            update.message.reply_text(f"✅ Остатки откатились на {chosen_date}", reply_markup=get_menu_keyboard(update.effective_user.id, menu="assembly"))

            # Отчёт
            df_report = df[["Артикул", "Остаток"]].rename(columns={"Остаток": "Стало"})
            df_report.insert(1, "+/-", "Откат")
            df_report.insert(1, "Было", "-")
            username = update.effective_user.full_name or update.effective_user.username
            ts = datetime.now().strftime("%d.%m.%Y %H:%M")
            save_and_send_excel(
                df_report, "rollback_report",
                f"♻️ Откат остатков\n👤 {username}\n🕓 {ts}",
                update, context
            )
        else:
            update.message.reply_text("⚠️ Не удалось применить откат: артикулы не найдены.")

    except Exception as e:
        logger.exception("perform_stock_rollback")
        update.message.reply_text(f"❌ Ошибка при выполнении отката:\n{e}", reply_markup=get_menu_keyboard(update.effective_user.id, menu="assembly"))

# ——— СОХРАНЕНИЕ EXCEL И ОТПРАВКА ————————————————————

def save_stock_snapshot(user, reason=""):
        try:
            vals = sheet_stock.get_all_values()
            if len(vals) < 2:
                return

            hdr = vals[0]
            art_i = hdr.index("Артикул")
            qty_i = hdr.index("Остатки склад")
            now = datetime.now()
            username = user.full_name or user.username or f"user_id {user.id}"
            rows = []

            for row in vals[1:]:
                art = str(row[art_i]).strip()
                qty = to_int(row[qty_i])
                rows.append([
                    art, qty,
                    now.strftime("%d.%m.%Y %H:%M"),
                    username,
                    reason
                ])

            try:
                sheet_history = sh.worksheet(SHEET_STOCK_HISTORY)
            except:
                sheet_history = sh.add_worksheet(title=SHEET_STOCK_HISTORY, rows=1000, cols=5)
                sheet_history.append_row(["Артикул", "Остаток", "Дата", "Пользователь", "Причина"], value_input_option="USER_ENTERED")

            # Проверка количества строк: если >= 1000, удалим самую старую дату
            existing = sheet_history.get_all_values()
            if len(existing) >= 1000:
                df_existing = pd.DataFrame(existing[1:], columns=existing[0])
                if not df_existing.empty and "Дата" in df_existing.columns:
                    oldest_date = df_existing["Дата"].sort_values().iloc[0]
                    df_trimmed = df_existing[df_existing["Дата"] != oldest_date]

                    # Очистим и перезапишем данные
                    sheet_history.clear()
                    sheet_history.append_row(["Артикул", "Остаток", "Дата", "Пользователь", "Причина"], value_input_option="USER_ENTERED")
                    sheet_history.append_rows(df_trimmed.values.tolist(), value_input_option="USER_ENTERED")

            # Добавим новые строки
            sheet_history.append_rows(rows, value_input_option="USER_ENTERED")

        except Exception:
            logger.exception("Ошибка при сохранении снимка остатков")

# ДОБАВЬ в начало файла (после save_stock_snapshot)
def rollback_to_last_snapshot(user):
    try:
        sheet_history = sh.worksheet(SHEET_STOCK_HISTORY)
        data = sheet_history.get_all_values()
        if len(data) < 2:
            return False

        hdr = data[0]
        df = pd.DataFrame(data[1:], columns=hdr)

        # Отфильтруем по пользователю и причине
        username = user.full_name or user.username or f"user_id {user.id}"
        df_user = df[(df["Пользователь"] == username) & (df["Причина"] == "pre_order_update")]

        if df_user.empty:
            return False

        # Найдём самую последнюю дату
        last_date = df_user["Дата"].max()
        df_latest = df_user[df_user["Дата"] == last_date]

        stock_vals = sheet_stock.get_all_values()
        hdr_stock = stock_vals[0]
        art_i = hdr_stock.index("Артикул")
        qty_i = hdr_stock.index("Остатки склад")
        art_row_map = {r[art_i]: i + 2 for i, r in enumerate(stock_vals[1:])}

        updates = []
        for row in df_latest.itertuples(index=False):
            art = getattr(row, "Артикул")
            qty = int(getattr(row, "Остаток"))
            if art in art_row_map:
                updates.append({
                    "range": rowcol_to_a1(art_row_map[art], qty_i + 1),
                    "values": [[qty]]
                })

        if updates:
            sheet_stock.batch_update(updates)
            return True

        return False

    except Exception:
        logger.exception("Ошибка при откате к последнему снимку")
        return False

def save_and_send_excel(df, prefix, caption, update, context):
    from io import BytesIO
    from telegram import InputFile

    bio = BytesIO()
    bio.name = f"{prefix}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    df.to_excel(bio, index=False)
    bio.seek(0)

    context.bot.send_document(
        chat_id=update.effective_chat.id,
        document=InputFile(bio),
        filename=bio.name,
        caption=caption,
        reply_markup=get_menu_keyboard(update.effective_user.id, menu="assembly")
    )

@queue_guard
def process_orders(update: Update, context: CallbackContext):
    try:
        logger.info("➡️ process_orders() вызвана")

        def send_reply(text):
            if update.message:
                update.message.reply_text(
                    text,
                    reply_markup=get_menu_keyboard(update.effective_user.id, menu="assembly")
                )
            else:
                update.callback_query.edit_message_text(
                    text,
                    reply_markup=get_menu_keyboard(update.effective_user.id, menu="assembly")
                )

        CHAT_ID = update.effective_chat.id
        user = update.effective_user
        username = user.full_name or user.username or f"user_id {user.id}"
        ts = datetime.now().strftime("%d.%m.%Y %H:%M")

        # 1. Загружаем заказы
        try:
            orders_raw = sheet_orders.get_all_values()
            logger.info("✅ Загружен лист 'Заказы'")
        except Exception:
            logger.exception("❌ Не удалось прочитать лист 'Заказы'")
            send_reply("❌ Не удалось прочитать лист 'Заказы'")
            return

        if len(orders_raw) < 2:
            logger.info("⚠️ Лист 'Заказы' пуст")
            send_reply("❗️ Лист «Заказы» пуст.")
            return

        headers, rows = orders_raw[0], orders_raw[1:]
        hdr_map = {h.strip().lower(): i for i, h in enumerate(headers)}
        col_art = hdr_map.get("артикул")
        col_qty = hdr_map.get("количество")
        col_mp  = hdr_map.get("маркетплейс")

        if col_art is None or col_mp is None:
            logger.error("❌ Не найдены колонки 'Артикул' и 'Маркетплейс'")
            send_reply("❌ Не найдены колонки 'Артикул' и 'Маркетплейс'")
            return

        orders = []
        for r in rows:
            try:
                art = str(r[col_art]).strip()
                qty = to_int(r[col_qty], 1) if col_qty is not None else 1
                mp  = str(r[col_mp]).strip()
                if art:
                    orders.append((art, qty, mp))
            except Exception as e:
                logger.warning(f"⚠️ Ошибка чтения строки заказа {r}: {e}")

        if not orders:
            logger.info("⚠️ Нет валидных заказов")
            send_reply("❗️ Нет валидных строк в заказах.")
            return

        logger.info(f"📦 Найдено {len(orders)} заказов")

        # 2. Загружаем остатки и справочники
        stock_vals = sheet_stock.get_all_values()
        stock_hdr  = stock_vals[0]
        stock_rows = stock_vals[1:]
        logger.info("✅ Остатки с листа 'Наш склад' загружены")

        main_data = sheet_main.get_all_records()
        logger.info("✅ Основные данные загружены")

        try:
            sheet_kits   = sh.worksheet("Комплекты")
            kits_data    = sheet_kits.get_all_records(expected_headers=[
                "Артикул","Название","Яч","Состав","Наименование","Кол-во",
                "Себес","Склад","Остаток поставщик","Сборка","Заказать","Ячейка"
            ])
        except Exception as e:
            kits_data = []
            logger.warning(f"⚠️ Не удалось загрузить лист 'Комплекты': {e}")

        # 3. Формируем словари
        kits_map = {}
        for row in kits_data:
            kit_art  = str(row.get("Артикул","")).strip()
            comp_art = str(row.get("Состав","")).strip()
            qty      = to_int(row.get("Кол-во",1))
            if kit_art and comp_art:
                kits_map.setdefault(kit_art,[]).append({"Артикул": comp_art, "Кол-во": qty})

        extra_main = {
            str(r.get("Артикул","")).strip(): {
                "Название":  str(r.get("Название","")).strip(),
                "Упаковка":  str(r.get("Упаковка","") or r.get("упаковка","")).strip(),
                "Бренд":     str(r.get("Бренд","")).strip(),
                "Опт":       to_int(r.get("Опт"))
            }
            for r in main_data
        }

        art_i  = stock_hdr.index("Артикул")
        qty_i  = stock_hdr.index("Остатки склад")
        cell_i = stock_hdr.index("Ячейка") if "Ячейка" in stock_hdr else None
        pack_i = stock_hdr.index("Упаковка") if "Упаковка" in stock_hdr else None

        initial_stock_map = {
            str(row[art_i]).strip(): to_int(row[qty_i])
            for row in stock_rows
            if art_i < len(row) and qty_i < len(row)
        }
        stock_map = initial_stock_map.copy()

        extra_info = {
            str(row[art_i]).strip(): {
                "Ячейка":   (row[cell_i] if cell_i is not None and cell_i < len(row) else ""),
                "Упаковка": (row[pack_i] if pack_i is not None and pack_i < len(row) else "")
            }
            for row in stock_rows
        }

        # 4. Предснапшот остатков
        save_stock_snapshot(user, reason="pre_order_update")

        move_recs = []
        supply_recs = []
        warnings    = []
        kits_assembled = 0

        # 5. Основной цикл по заказам
        for art, q, mp in orders:
            # 5.1. Если готовый комплект есть на складе
            if stock_map.get(art,0) >= q:
                stock_map[art] -= q
                move_recs.append({
                    "Название":     extra_main.get(art,{}).get("Название",art),
                    "Артикул":      art,
                    "Комплект":     "",
                    "Кол-во":       q,
                    "Мп":           mp,
                    "Остатки":      stock_map[art],
                    "Упаковка":     extra_main.get(art,{}).get("Упаковка",""),
                    "Ячейка":       extra_info.get(art,{}).get("Ячейка",""),
                    "Себестоимость":extra_main.get(art,{}).get("Опт",0)
                })
                continue

            # 5.2. Если это позиция-комплект
            if art in kits_map:
                kit_brand = extra_main.get(art,{}).get("Бренд","").lower().strip()
                missing, has_all = [], True

                # проверяем компоненты
                for comp in kits_map[art]:
                    c_art, c_qty = comp["Артикул"], comp["Кол-во"]*q
                    if stock_map.get(c_art,0) < c_qty:
                        has_all = False
                        missing.append(c_art)

                #  Гладиатор — на отсутствие реагируем только предупреждением
                if kit_brand=="гладиатор" and not has_all:
                    warnings.append(
                        f"🛑 Комплект {art} (Гладиатор): не хватает компонентов ({', '.join(missing)})"
                    )
                    continue

                # собираем полностью
                if has_all:
                    for comp in kits_map[art]:
                        c_art, c_qty = comp["Артикул"], comp["Кол-во"]*q
                        stock_map[c_art] -= c_qty
                    move_recs.append({
                        "Название":     extra_main.get(art,{}).get("Название",art),
                        "Артикул":      art,
                        "Комплект":     "",
                        "Кол-во":       q,
                        "Мп":           mp,
                        "Остатки":      0,
                        "Упаковка":     extra_main.get(art,{}).get("Упаковка",""),
                        "Ячейка":       "",
                        "Себестоимость":extra_main.get(art,{}).get("Опт",0)
                    })
                    kits_assembled += 1
                    warnings.append(f"🧩 Комплект {art} собран из составляющих")
                    continue

                # частично: вычитаем, остальное — в заказ поставщику/предупреждение
                for comp in kits_map[art]:
                    c_art, c_qty = comp["Артикул"], comp["Кол-во"]*q
                    avail = stock_map.get(c_art,0)
                    br   = extra_main.get(c_art,{}).get("Бренд","")
                    if avail >= c_qty:
                        stock_map[c_art] -= c_qty
                        move_recs.append({
                            "Название": extra_main[c_art]["Название"],
                            "Артикул":  c_art,
                            "Комплект": art,
                            "Кол-во":   c_qty,
                            "Мп":       mp,
                            "Остатки":  stock_map[c_art],
                            "Упаковка": extra_info[c_art]["Упаковка"],
                            "Ячейка":   extra_info[c_art]["Ячейка"],
                            "Себестоимость": extra_main[c_art]["Опт"]
                        })
                    else:
                        if br in TECHNOMARIN_BRANDS:
                            supply_recs.append([c_art, c_qty])
                            warnings.append(f"⚠️ Комплект {art}: {c_art} не хватает, заказ поставщику")
                        else:
                            warnings.append(f"❗ Комплект {art}: {c_art} отсутствует в остатках")
                continue

            # 5.3. Обычная позиция
            br = extra_main.get(art,{}).get("Бренд","")
            name = extra_main.get(art,{}).get("Название","")
            pack = extra_info.get(art,{}).get("Упаковка","")
            cell = extra_info.get(art,{}).get("Ячейка","")
            opt  = extra_main.get(art,{}).get("Опт",0)
            avail = stock_map.get(art)

            if avail is not None:
                take = min(avail, q)
                if take>0:
                    stock_map[art] -= take
                    move_recs.append({
                        "Название":     name,
                        "Артикул":      art,
                        "Комплект":     "",
                        "Кол-во":       take,
                        "Мп":           mp,
                        "Остатки":      stock_map[art],
                        "Упаковка":     pack,
                        "Ячейка":       cell,
                        "Себестоимость":opt
                    })
                lack = q - take
                if lack>0:
                    if br in TECHNOMARIN_BRANDS:
                        supply_recs.append([art, lack])
                        if take>0:
                            warnings.append(f"⚠️ Проверьте количество: {art} (ячейка {cell})")
                    else:
                        move_recs.append({
                            "Название":     name,
                            "Артикул":      art,
                            "Комплект":     "",
                            "Кол-во":       lack,
                            "Мп":           mp,
                            "Остатки":      0,
                            "Упаковка":     pack,
                            "Ячейка":       cell,
                            "Себестоимость":opt
                        })
                        warnings.append(f"❗ Товар {art} (ячейка {cell}) отсутствует в остатках")
            else:
                # не было вообще в остатках
                if br in TECHNOMARIN_BRANDS:
                    supply_recs.append([art, q])
                else:
                    move_recs.append({
                        "Название":     name,
                        "Артикул":      art,
                        "Комплект":     "",
                        "Кол-во":       q,
                        "Мп":           mp,
                        "Остатки":      0,
                        "Упаковка":     pack,
                        "Ячейка":       cell,
                        "Себестоимость":opt
                    })
                    warnings.append(f"❗ Товар {art} отсутствует в остатках")

        # 6. Запись и отправка Перемещения
        if move_recs:
            directly_ordered = {art for art,_,_ in orders}
            cleaned = []
            for rec in move_recs:
                if rec["Комплект"] and rec["Артикул"] not in directly_ordered:
                    continue
                cleaned.append(rec)

            df_move = pd.DataFrame(cleaned)[MOVE_COLUMNS].sort_values(["Ячейка","Мп"])
            try:
                sheet_move = sh.worksheet(SHEET_MOVE_NAME)
            except:
                sheet_move = sh.add_worksheet(title=SHEET_MOVE_NAME, rows=1000, cols=15)

            sheet_move.batch_clear([f"A2:{chr(64+len(MOVE_COLUMNS))}1000"])
            sheet_move.append_row(MOVE_COLUMNS, value_input_option="USER_ENTERED")
            sheet_move.append_rows(df_move.values.tolist(), value_input_option="USER_ENTERED")
            sheet_move.update_acell("H1","=SUM(H2:H1000)")

            # обновление себестоимости
            updates = []
            for idx,rec in df_move.iterrows():
                updates.append({
                    "range": f"H{idx+2}",
                    "values": [[extra_main.get(rec["Артикул"],{}).get("Опт",0)]]
                })
            if updates:
                sheet_move.batch_update(updates)

            # формируем и отправляем .xlsx
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.styles import Border, Side, numbers
            from io import BytesIO

            wb = Workbook(); ws = wb.active; ws.title="Перемещение"
            for r in dataframe_to_rows(df_move,index=False,header=True):
                ws.append(r)
            ws.auto_filter.ref = ws.dimensions
            thin = Border(Side('thin'),Side('thin'),Side('thin'),Side('thin'))
            for row in ws.iter_rows(min_row=1,max_row=ws.max_row,max_col=ws.max_column):
                for cell in row:
                    cell.border = thin
            for col in ws.iter_cols(min_col=1,max_col=ws.max_column):
                if col[0].value=="Ячейка":
                    for c in col[1:]:
                        c.number_format="@"
            for row in ws.iter_rows(min_row=2,max_col=len(MOVE_COLUMNS)):
                for c in row:
                    if isinstance(c.value,(int,float)):
                        c.number_format=numbers.FORMAT_NUMBER

            bio = BytesIO(); wb.save(bio); bio.seek(0)
            bio.name = f"Перемещение_{datetime.now():%Y%m%d_%H%M%S}.xlsx"

            if not context.user_data.get("move_sent"):
                context.user_data["move_sent"] = True
                context.bot.send_document(
                    chat_id=CHAT_ID,
                    document=InputFile(bio),
                    filename=bio.name,
                    caption="📦 Файл Перемещение",
                    reply_markup=get_menu_keyboard(update.effective_user.id, menu="assembly")
                )
            else:
                logger.warning("⚠️ Повторная попытка отправки Перемещения пропущена")

        # 7. Пост-снапшот и отчёт об изменениях
        save_stock_snapshot(user, reason="post_order_update")
        report_text, report_file = compare_stock_snapshots(user)
        if report_file:
            with open(report_file,"rb") as f:
                context.bot.send_document(
                    chat_id=CHAT_ID,
                    document=InputFile(f),
                    filename=os.path.basename(report_file),
                    caption="📉 Отчет по изменениям остатков",
                    reply_markup=get_menu_keyboard(update.effective_user.id, menu="assembly")
                )
            os.remove(report_file)
        if report_text:
            context.bot.send_message(
                chat_id=CHAT_ID,
                text=report_text,
                reply_markup=get_menu_keyboard(update.effective_user.id, menu="assembly")
            )

        # 8. Файл Заказ поставщику
        if supply_recs:
            from collections import defaultdict
            cnt = defaultdict(int)
            for art,qty in supply_recs:
                cnt[art]+=qty
            df_supply = pd.DataFrame([{"Арт":a,"Кол":q} for a,q in cnt.items()])

            fn = f"Заказ_поставщику_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
            df_supply.to_excel(fn,index=False)

            with open(fn,"rb") as f:
                context.bot.send_document(
                    chat_id=CHAT_ID,
                    document=InputFile(f),
                    filename=fn,
                    caption="🛒 Файл Заказ поставщику",
                    reply_markup=get_menu_keyboard(update.effective_user.id, menu="assembly")
                )
            os.remove(fn)
            logger.info("✅ Заказ поставщику сформирован")

        # 9. Inline-проверка актуальности «было→стало»
        summary = f"✅ Перемещение: {len(move_recs)} позиций\n✅ Заказ ТМ: {len(supply_recs)}"
        if kits_assembled:
            summary += f"\n🧩 Собрано комплектов: {kits_assembled}"
        if warnings:
            summary += "\n\n⚠️ Внимание:\n" + "\n".join(warnings)

        check_actual = []
        for art,_,_ in orders:
            old = initial_stock_map.get(art,0)
            new = stock_map.get(art,0)
            if old>0 and new==0 and extra_main.get(art,{}).get("Бренд","") in TECHNOMARIN_BRANDS:
                check_actual.append((art, extra_info[art]["Ячейка"], old, new))

        if check_actual:
            kb = InlineKeyboardMarkup([[ 
                InlineKeyboardButton("Да", callback_data="actual_yes"),
                InlineKeyboardButton("Нет", callback_data="actual_no")
            ]])
            msg = "❓ Проверьте актуальность позиций:\n"
            for art,cell,old,new in check_actual:
                msg += f"- {art} (ячейка {cell}): было {old} → стало {new}\n"
            context.bot.send_message(chat_id=CHAT_ID, text=msg, reply_markup=kb)

        # 10. Финальный summary и очистка
        context.bot.send_message(
            chat_id=CHAT_ID,
            text=summary,
            reply_markup=get_menu_keyboard(update.effective_user.id, menu="assembly")
        )
        sheet_orders.clear()
        sheet_orders.append_row(
            ["Артикул","Количество","Маркетплейс"],
            value_input_option="USER_ENTERED"
        )
        logger.info("✅ Заказы очищены, завершение")

    except Exception as e:
        logger.exception("❌ Ошибка в process_orders")
        restored = rollback_to_last_snapshot(user)
        if restored:
            context.bot.send_message(
                chat_id=CHAT_ID,
                text="♻️ Остатки восстановлены после ошибки",
                reply_markup=get_menu_keyboard(update.effective_user.id, menu="assembly")
            )
        send_reply(f"❌ Ошибка при обработке заказов:\n{e}")

def start(update: Update, context: CallbackContext):
    context.user_data.clear()
    context.user_data["menu"] = "main"  # ✅ Явно указываем, что мы в главном меню
    update.message.reply_text(
        "📊 *Меню бота управления остатками*\nВыберите действие:",
        parse_mode="Markdown",
        reply_markup=get_menu_keyboard(update.effective_user.id, menu="main")  # ✅ Главное меню
    )

def save_stock_snapshot(user, reason=""):
    try:
        vals = sheet_stock.get_all_values()
        if len(vals) < 2:
            return

        hdr = vals[0]
        art_i = hdr.index("Артикул")
        qty_i = hdr.index("Остатки склад")
        now = datetime.now()
        username = user.full_name or user.username or f"user_id {user.id}"
        rows = []

        for row in vals[1:]:
            art = str(row[art_i]).strip()
            qty = to_int(row[qty_i])
            rows.append([
                art, qty,
                now.strftime("%d.%m.%Y %H:%M"),
                username,
                reason
            ])

        try:
            sheet_history = sh.worksheet(SHEET_STOCK_HISTORY)
        except:
            sheet_history = sh.add_worksheet(title=SHEET_STOCK_HISTORY, rows=1000, cols=5)
            sheet_history.append_row(["Артикул", "Остаток", "Дата", "Пользователь", "Причина"], value_input_option="USER_ENTERED")

        sheet_history.append_rows(rows, value_input_option="USER_ENTERED")

    except Exception:
        logger.exception("Ошибка при сохранении снимка остатков")

    # 🔁 Вставь эту функцию рядом с save_stock_snapshot (если ещё нет)
def compare_stock_snapshots(user):
    try:
        sheet = sh.worksheet(SHEET_STOCK_HISTORY)
        records = sheet.get_all_records()
        if len(records) < 2:
            return None, None

        last_ts = records[-1]["Дата"]
        prev_ts = None

        # Находим последний snapshot до текущего
        for r in reversed(records[:-1]):
            if r["Причина"].startswith("pre_order"):
                prev_ts = r["Дата"]
                break

        if not prev_ts:
            return None, None

        # Отбираем данные по 2 временным меткам
        prev_snapshot = [r for r in records if r["Дата"] == prev_ts]
        last_snapshot = [r for r in records if r["Дата"] == last_ts]

        # Формируем словари
        prev_map = {r["Артикул"]: int(r["Остаток"]) for r in prev_snapshot}
        last_map = {r["Артикул"]: int(r["Остаток"]) for r in last_snapshot}

        rows = []
        text_lines = []
        for art in sorted(set(map(str, prev_map.keys())).union(map(str, last_map.keys()))):
            before = prev_map.get(art, 0)
            after = last_map.get(art, 0)
            diff = after - before
            if diff != 0:
                rows.append([art, before, after, diff])
                text_lines.append(f"{art}: {before} → {after} ({'+' if diff > 0 else ''}{diff})")

        if not rows:
            return None, None

        df = pd.DataFrame(rows, columns=["Артикул", "Было", "Стало", "Изменение"])
        path = f"Отчет_изменения_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        df.to_excel(path, index=False)

        username = user.full_name or user.username or f"user_id {user.id}"
        header = f"📋 Изменения остатков ({username}):\n\n"
        return header + "\n".join(text_lines[:20]), path

    except Exception:
        logger.exception("compare_stock_snapshots:")
        return None, None

@queue_guard
def handle_callback(update: Update, context: CallbackContext):
    query = update.callback_query
    query.answer()
    data = query.data
    CHAT_ID = query.message.chat.id

    if data == "actual_no":
        # Пользователь указал, что не все позиции актуальны
        check_actual_recs = context.user_data.get("check_actual", [])
        art_list = [art for art, _ in check_actual_recs]
        context.user_data["awaiting_missing_actual"] = art_list

        msg = (
            "🛠️ Пришлите артикулы, которые оказались НЕ в наличии,\n"
            "в одном сообщении, через пробел или с новой строки.\n\n"
            "Полный список:\n" + "\n".join(f"- {art}" for art in art_list)
        )
        context.bot.send_message(chat_id=CHAT_ID, text=msg, reply_markup=get_menu_keyboard(update.effective_user.id, menu="assembly"))
        query.edit_message_text("⏳ Ожидаю список неактуальных артикулов…")

    elif data == "actual_yes":
        context.user_data.pop("check_actual", None)
        context.bot.send_message(chat_id=CHAT_ID, text="✅ Хорошо, позиции актуальны", reply_markup=get_menu_keyboard(update.effective_user.id, menu="assembly"))
        query.edit_message_text("✅ Подтверждено")

def handle_missing_actual(update: Update, context: CallbackContext):
    text = update.message.text.strip()
    if not text:
        update.message.reply_text("⚠️ Пожалуйста, пришлите хотя бы один артикул.")
        return

    art_raw = text.replace(",", " ").replace(";", " ")
    arts = set(art.strip() for art in art_raw.split() if art.strip())
    allowed = set(context.user_data.pop("awaiting_missing_actual", []))
    arts = [a for a in arts if a in allowed]

    if not arts:
        update.message.reply_text("⚠️ Ни один артикул не найден среди запрашиваемых.")
        return

    extra_main = context.bot_data.get("extra_main", {})
    now = datetime.now().strftime("%d.%m.%Y")
    tm_rows = []
    for art in arts:
        name = extra_main.get(art, {}).get("Название", "")
        brand = extra_main.get(art, {}).get("Бренд", "")
        tm_rows.append([art, name, 1, brand, now])

    try:
        sheet_tm = sh.worksheet(SHEET_TM_NAME)
    except:
        sheet_tm = sh.add_worksheet(title=SHEET_TM_NAME, rows=1000, cols=10)
        sheet_tm.append_row(["Артикул", "Название", "Количество", "Бренд", "Дата"], value_input_option="USER_ENTERED")

    sheet_tm.append_rows(tm_rows, value_input_option="USER_ENTERED")

    df_supply = pd.DataFrame([{"Арт": art, "Кол": 1} for art in arts])
    fn = f"Заказ_поставщику_неактуальные_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    df_supply.to_excel(fn, index=False)

    with open(fn, "rb") as f:
        context.bot.send_document(
            chat_id=update.effective_chat.id,
            document=InputFile(f),
            filename=fn,
            caption="🛒 Добавлены неактуальные позиции в заказ",
            reply_markup=get_menu_keyboard(update.effective_user.id, menu="assembly")
        )
    os.remove(fn)
    update.message.reply_text("📥 Заказ по неактуальным позициям сформирован.", reply_markup=get_menu_keyboard(update.effective_user.id, menu="assembly"))


def main():
        updater = Updater("7813534646:AAFwM01aa3Nk7bfzzskPU4iCFFqr9cm1v7o", use_context=True)
        dp = updater.dispatcher

        # — Команды
        dp.add_handler(CommandHandler("start", start))

        # — Callback-кнопки (inline-кнопки "Актуальны — да/нет")
        dp.add_handler(CallbackQueryHandler(handle_callback))

        # — Документы Excel / CSV
        dp.add_handler(MessageHandler(TgFilters.document, handle_document))

        # — Текстовые команды (reply-кнопки, ручной ввод)
        dp.add_handler(MessageHandler(TgFilters.text & ~TgFilters.command, handle_message))

        # — Запуск
        updater.start_polling()
        logger.info("✅ Бот запущен")
        updater.idle()

if __name__ == "__main__":
    main()
